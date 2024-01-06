from tkinter import *
import datetime
from tkinter import ttk 
import openpyxl
from openpyxl import Workbook

root = Tk()
root.geometry('950x552+150+150')
root.iconbitmap('img/store-64.ico')
root.title('Market tools for Building')
now = datetime.datetime.now()
date=now.strftime("%Y-%m-%d")
time=now.strftime("%I:%M %p")
# ============ EXCEL ================
wb =Workbook()
ws = wb.active
ws.title = 'customer'
ws["A1"] = 'Full Name'
ws["B1"] = 'Phone'
ws["C1"] = 'Address'
ws["D1"] = 'Total'
ws["E1"] = 'Date'
ws["F1"] = 'Time'
wb.save('Custemor.xlsx')

# ============ Price ================
menu = {
    0:['Hammer',20]
    ,1:['Drill',30]
    ,2:['Saw',18],
    3:['Mixer',30],
    4:['Welder',30],
    5:['malte',5],
    6:['Plumber',30],
    7:['rabsh',15],
    8:['BOOM',3],
    9:['wood',1],
    10:['Carpenter',300],
    11:['Muddy',320]
}
def bill():
    global E_name
    global E_phone
    global E_address
    global E_total
    global E_date
    global E_time

    root.geometry('1205x552')
    f4=Frame(root,bg='#5f7161',width=250,height=434,bd=2,relief=GROOVE)
    f4.place(x=950,y=1)
    
    L_name = Label(f4,text="Name : ",bg='#5f7161',fg='white')
    L_phone =Label(f4,text="Phone : ",bg='#5f7161',fg='white')
    L_address=Label(f4,text="Address : ",bg='#5f7161',fg='white')
    L_total=Label(f4,text="Total : ",bg='#5f7161',fg='white')
    L_Date=Label(f4,text="Date : ",bg='#5f7161',fg='white')
    L_time=Label(f4,text="Date : ",bg='#5f7161',fg='white')
    L_imge=Label(f4,image= image_menu13,bg='#5f7161')

    L_name.place(x=0,y=10)
    L_phone.place(x=0,y=40)
    L_address.place(x=0,y=70)
    L_total.place(x=0,y=100)
    L_Date.place(x=0,y=130)
    L_time.place(x=0,y=160)
    L_imge.place(x=110,y=320)

    E_name =Entry(f4,width=20,font=font1,justify=CENTER,bg='#5f7161',fg='white')
    E_phone =Entry(f4,width=20,font=font1,justify=CENTER,bg='#5f7161',fg='white')
    E_address=Entry(f4,width=20,font=font1,justify=CENTER,bg='#5f7161',fg='white')
    E_total=Entry(f4,width=20,font=font1,justify=CENTER,bg='#5f7161',fg='white')
    E_date=Entry(f4,width=20,font=font1,justify=CENTER,bg='#5f7161',fg='white')
    E_time=Entry(f4,width=20,font=font1,justify=CENTER,bg='#5f7161',fg='white')

    E_name.place(x=45,y=10)
    E_phone.place(x=45,y=40)
    E_address.place(x=45,y=70)
    E_total.place(x=45,y=100)
    E_date.place(x=45,y=130)
    E_time.place(x=45,y=160)
    
    
    add=Button(f4,text='ADD',width=31,cursor='hand2',bg='#eddbc0',command=save)
    clear=Button(f4,text='Clear',width=31,cursor='hand2',bg='#eddbc0',command=E_clear)
    search=Button(f4,text='Search',width=31,cursor='hand2',bg='#eddbc0',command=S_search)
    delet=Button(f4,text='Delet',width=31,cursor='hand2',bg='#eddbc0',command=D_cust)
    add.place(x=12,y=190)
    clear.place(x=12,y=220)
    search.place(x=12,y=250)
    delet.place(x=12,y=280)
    


    

    total = 0

    for item in trv.get_children():
        trv.delete(item)
    for i in range(len(sb)):
        if(int(sb[i].get())>0):
            price =int(sb[i].get())*menu[i][1]
            total += price
            myst=(str(menu[i][1]),str(sb[i].get()),str(price))
            trv.insert('', 'end',iid=i,text=menu[i][0], values=myst)
    finall = total
    E_total.insert('1',str(finall) + '$')
    E_date.insert('1',str(date))
    E_time.insert('1',str(time))
def clear():
    for item in trv.get_children():
        trv.delete(item)
    E_name.delete('0',END)
    E_phone.delete('0',END)
    E_address.delete('0',END)
    E_total.delete('0',END)
    E_date.delete('0',END)
    E_time.delete('0',END)
def E_clear():
    E_name.delete('0',END)
    E_phone.delete('0',END)
    E_address.delete('0',END)
    E_total.delete('0',END)
    E_date.delete('0',END)
    E_time.delete('0',END)
def save():
    name = E_name.get()
    phone = E_phone.get()
    address = E_address.get()
    total = E_total.get()
    date = E_date.get()
    time = E_time.get()
    excel=openpyxl.load_workbook('Custemor.xlsx')
   
    file = excel.active
   
    file.cell(column=1,row=file.max_row+1,value=name)
    file.cell(column=2,row=file.max_row,value=phone)
    file.cell(column=3,row=file.max_row,value=address)
    file.cell(column=4,row=file.max_row,value=total)
    file.cell(column=5,row=file.max_row,value=date)
    file.cell(column=6,row=file.max_row,value=time)
    excel.save('Custemor.xlsx')
# === fill phone by search name in E_phone
def S_search():
    name = E_name.get()
    excel=openpyxl.load_workbook('Custemor.xlsx')
    file = excel.active
    for i in range(file.max_row):
        if(file.cell(column=1,row=i+1).value==name):
            E_phone.insert('0',file.cell(column=2,row=i+1).value)
            E_address.insert('0',file.cell(column=3,row=i+1).value)
            E_total.insert('0',file.cell(column=4,row=i+1).value)
            E_date.insert('0',file.cell(column=5,row=i+1).value)
            E_time.insert('0',file.cell(column=6,row=i+1).value)
            break
# ====== delet  from Custemor.xlsx  name phone address Total date time by name ======
def D_cust():
    name = E_name.get()
    excel=openpyxl.load_workbook('Custemor.xlsx')
    file = excel.active
    for i in range(file.max_row):
        if(file.cell(column=1,row=i+1).value==name):
            E_name.delete('0',END)
            E_phone.delete('0',END)
            E_address.delete('0',END)
            E_total.delete('0',END)
            E_date.delete('0',END)
            E_time.delete('0',END)
            file.delete_rows(i+1)
            excel.save('Custemor.xlsx')
            break
# ============ FRAMES ===============
# ============ frame 1 ==============
f1 = Frame(root,bg='silver')
f1.place(x=1, y=1, width=600, height=550)
# ============ frame 2 ==============
f2 = Frame(root,bg='gray')
f2.place(x=610, y=1, width=343, height=550)

trv = ttk.Treeview(f2,selectmod='browse')

trv.place(x=0, y=0, width=340, height=550)
trv["columns"]=(1,2,3)
trv.column("0",width=10,anchor='c')

trv.column("1",width=10,anchor='center')
trv.column("2",width=10,anchor='center')
trv.column("3",width=10,anchor='center')

trv.heading("#0", text="Building")
trv.heading("#1", text="Price")
trv.heading("#2", text="Quantity")
trv.heading("#3", text="Total")

# ============ LABEL ================
title = Label(f1,text='Building equipment sale project',font=('Perpetua',16),fg='white',bg='#5f7161')# Lucida Calligraphy italic
title.pack(fill=X)
# ============ IMAGES ===============
image_menu1 =  image=PhotoImage(file='img/icons8-hammer-60.png')
image_menu2 =  image=PhotoImage(file='img/icons8-saw-32.png')
image_menu3 =  image=PhotoImage(file='img/icons8-wagon-32.png')
image_menu4 =  image=PhotoImage(file='img/icons8-broom-64.png')
image_menu5 =  image=PhotoImage(file='img/icons8-hammer-60.png') 
image_menu6 =  image=PhotoImage(file='img/icons8-saw-32.png')
image_menu7 =  image=PhotoImage(file='img/icons8-wagon-32.png')
image_menu8 =  image=PhotoImage(file='img/icons8-broom-64.png')
image_menu9 =  image=PhotoImage(file='img/icons8-hammer-60.png')
image_menu10 =  image=PhotoImage(file='img/icons8-saw-32.png')
image_menu11 =  image=PhotoImage(file='img/icons8-wagon-32.png') 
image_menu12 =  image=PhotoImage(file='img/icons8-broom-64.png')
image_menu13 = image=PhotoImage(file='img/icons8-tellonym-50.png')

#========== BUTONNS =================
button_menu1 = Button(f1,bg='#efead8',bd=1,relief=SOLID,cursor='hand2',width=88,height=85,image=image_menu1,text='Hammer',compound=TOP)
button_menu2 = Button(f1,bg='#efead8',bd=1,relief=SOLID,cursor='hand2',width=88,height=85,image=image_menu2,text='Saw',compound=TOP)
button_menu3 = Button(f1,bg='#efead8',bd=1,relief=SOLID,cursor='hand2',width=88,height=85,image=image_menu3,text='Wagon',compound=TOP)
button_menu4 = Button(f1,bg='#efead8',bd=1,relief=SOLID,cursor='hand2',width=88,height=85,image=image_menu4,text='Broom',compound=TOP)
button_menu5 = Button(f1,bg='#efead8',bd=1,relief=SOLID,cursor='hand2',width=88,height=85,image=image_menu5,text='Hammer',compound=TOP)
button_menu6 = Button(f1,bg='#efead8',bd=1,relief=SOLID,cursor='hand2',width=88,height=85,image=image_menu6,text='Saw',compound=TOP)
button_menu7 = Button(f1,bg='#efead8',bd=1,relief=SOLID,cursor='hand2',width=88,height=85,image=image_menu7,text='Wagon',compound=TOP)
button_menu8 = Button(f1,bg='#efead8',bd=1,relief=SOLID,cursor='hand2',width=88,height=85,image=image_menu8,text='Broom',compound=TOP)
button_menu9 = Button(f1,bg='#efead8',bd=1,relief=SOLID,cursor='hand2',width=88,height=85,image=image_menu9,text='Hammer',compound=TOP)
button_menu10 = Button(f1,bg='#efead8',bd=1,relief=SOLID,cursor='hand2',width=88,height=85,image=image_menu10,text='Saw',compound=TOP)
button_menu11 = Button(f1,bg='#efead8',bd=1,relief=SOLID,cursor='hand2',width=88,height=85,image=image_menu11,text='Wagon',compound=TOP)
button_menu12 = Button(f1,bg='#efead8',bd=1,relief=SOLID,cursor='hand2',width=88,height=85,image=image_menu12,text='Broom',compound=TOP)
button_menu1.place(x=30,y=45)
button_menu2.place(x=170,y=45)
button_menu3.place(x=310,y=45)
button_menu4.place(x=450,y=45)
button_menu5.place(x=30,y=180)
button_menu6.place(x=170,y=180)
button_menu7.place(x=310,y=180)
button_menu8.place(x=450,y=180)
button_menu9.place(x=30,y=315)
button_menu10.place(x=170,y=315)
button_menu11.place(x=310,y=315)
button_menu12.place(x=450,y=315)
# ========= Varible + count ==============
sb = []
font1 =('Lucida Calligraphy italic',12,'normal')
sv1 =IntVar()
sv2 =IntVar()
sv3 =IntVar()
sv4 =IntVar()
sv5 =IntVar()
sv6 =IntVar()
sv7 =IntVar()
sv8 =IntVar()
sv9 =IntVar()
sv10 =IntVar()
sv11 =IntVar()
sv12 =IntVar()

sb1 = Spinbox(f1,from_=0,to_=5,font=font1,width=7,textvariable=sv1)
sb2 = Spinbox(f1,from_=0,to_=5,font=font1,width=7,textvariable=sv2)
sb3 = Spinbox(f1,from_=0,to_=5,font=font1,width=7,textvariable=sv3)
sb4 = Spinbox(f1,from_=0,to_=5,font=font1,width=7,textvariable=sv4)
sb5 = Spinbox(f1,from_=0,to_=5,font=font1,width=7,textvariable=sv5)
sb6 = Spinbox(f1,from_=0,to_=5,font=font1,width=7,textvariable=sv6)
sb7 = Spinbox(f1,from_=0,to_=5,font=font1,width=7,textvariable=sv7)
sb8 = Spinbox(f1,from_=0,to_=5,font=font1,width=7,textvariable=sv8)
sb9 = Spinbox(f1,from_=0,to_=5,font=font1,width=7,textvariable=sv9)
sb10 = Spinbox(f1,from_=0,to_=5,font=font1,width=7,textvariable=sv10)
sb11 = Spinbox(f1,from_=0,to_=5,font=font1,width=7,textvariable=sv11)
sb12 = Spinbox(f1,from_=0,to_=5,font=font1,width=7,textvariable=sv12)

sb1.place(x=30,y=140)
sb2.place(x=170,y=140)
sb3.place(x=310,y=140)
sb4.place(x=450,y=140)
sb5.place(x=30,y=275)
sb6.place(x=170,y=275)
sb7.place(x=310,y=275)
sb8.place(x=450,y=275)
sb9.place(x=30,y=410)
sb10.place(x=170,y=410)
sb11.place(x=310,y=410)
sb12.place(x=450,y=410)

sb.append(sb1)
sb.append(sb2)
sb.append(sb3)
sb.append(sb4)
sb.append(sb5)
sb.append(sb6)
sb.append(sb7)
sb.append(sb8)
sb.append(sb9)
sb.append(sb10)
sb.append(sb11)
sb.append(sb12)

B1 = Button(f1,text="buying",font=font1,bd=1,bg='#6d8b74',command=bill)
B2 = Button(f1,text="New bill",font=font1,bd=1,bg='#6d8b74',command=clear)
B3 = Button(f1,text="Rent materials",font=font1,bd=1,bg='#6d8b74')
B4 = Button(f1,text="Close the program",font=font1,bd=1,bg='#6d8b74',command=exit)

B1.place(x=0,y=480)
B2.place(x=110,y=480)
B3.place(x=250,y=480)
B4.place(x=420,y=480)







root.mainloop()